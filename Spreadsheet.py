from openpyxl import Workbook, load_workbook

# Writing data to Excel
wb = Workbook()
ws = wb.active
ws.title = "Student Data"

ws.append(["Name", "Age", "Marks"])
ws.append(["John", 20, 85])
ws.append(["Sara", 21, 90])
ws.append(["Lilly", 22, 95])

wb.save("students.xlsx")
print("Data written to students.xlsx successfully!")

# Reading data from Excel
print("\nReading data from students.xlsx:")
wb2 = load_workbook("students.xlsx")
ws2 = wb2.active

for row in ws2.iter_rows(values_only=True):
    print(row)
